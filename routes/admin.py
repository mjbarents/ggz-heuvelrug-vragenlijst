from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    session,
    flash,
    current_app,
    make_response,
)
from models import db, Question, Token, Response
from functools import wraps
from sqlalchemy import func
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("admin_logged_in"):
            if request.endpoint != "admin.admin_dashboard":
                flash(
                    "Log alstublieft in om toegang te krijgen tot het admin panel.",
                    "warning",
                )
            return redirect(url_for("admin.admin_login"))
        return f(*args, **kwargs)

    return decorated_function


@admin_bp.route("/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        if (
            username == current_app.config["ADMIN_USERNAME"]
            and password == current_app.config["ADMIN_PASSWORD"]
        ):
            session["admin_logged_in"] = True
            return redirect(url_for("admin.admin_dashboard"))
        else:
            flash("Ongeldige inloggegevens.", "error")

    return render_template("admin_login.html")


@admin_bp.route("/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    return redirect(url_for("admin.admin_login"))


@admin_bp.route("/")
@admin_required
def admin_dashboard():
    total_tokens = Token.query.count()
    used_tokens = Token.query.filter_by(used=True).count()
    total_responses = Response.query.count()
    total_questions = Question.query.count()

    responses = (
        db.session.query(
            Response.id,
            Response.token_id,
            Response.submitted_at,
            Question.text,
            Question.question_type,
            Response.score,
            Response.text_answer,
        )
        .join(Question)
        .order_by(Response.token_id, Question.order)
        .all()
    )

    grouped_responses = {}
    for response in responses:
        token_id = response.token_id
        if token_id not in grouped_responses:
            grouped_responses[token_id] = {
                "token_id": token_id,
                "submitted_at": response.submitted_at,
                "answers": [],
            }
        grouped_responses[token_id]["answers"].append(
            {
                "question": response.text,
                "question_type": response.question_type,
                "score": response.score,
                "text_answer": response.text_answer,
            }
        )

    questions = Question.query.order_by(Question.order).all()
    stats = []
    for question in questions:
        if question.question_type == "scale":
            scores = [
                r.score
                for r in Response.query.filter_by(question_id=question.id).all()
                if r.score is not None
            ]
            if scores:
                avg_score = sum(scores) / len(scores)
                stats.append(
                    {
                        "question": question.text,
                        "question_type": "scale",
                        "avg_score": round(avg_score, 1),
                        "responses": len(scores),
                        "scores": scores,
                    }
                )
        else:
            text_answers = [
                r.text_answer
                for r in Response.query.filter_by(question_id=question.id).all()
                if r.text_answer
            ]
            if text_answers:
                stats.append(
                    {
                        "question": question.text,
                        "question_type": "open",
                        "responses": len(text_answers),
                        "text_answers": text_answers,
                    }
                )

    return render_template(
        "admin_dashboard.html",
        total_tokens=total_tokens,
        used_tokens=used_tokens,
        total_responses=total_responses,
        total_questions=total_questions,
        grouped_responses=grouped_responses.values(),
        stats=stats,
    )


@admin_bp.route("/questions")
@admin_required
def admin_questions():
    questions = Question.query.order_by(Question.order).all()
    return render_template("admin_questions.html", questions=questions)


@admin_bp.route("/questions/add", methods=["POST"])
@admin_required
def admin_add_question():
    text = request.form.get("text")
    question_type = request.form.get("question_type", "scale")
    if question_type not in ("scale", "open"):
        question_type = "scale"
    if text:
        max_order = db.session.query(func.max(Question.order)).scalar() or 0
        question = Question(text=text, order=max_order + 1, question_type=question_type)
        db.session.add(question)
        db.session.commit()
    return redirect(url_for("admin.admin_questions"))


@admin_bp.route("/questions/edit/<int:id>", methods=["POST"])
@admin_required
def admin_edit_question(id):
    question = Question.query.get_or_404(id)
    text = request.form.get("text")
    question_type = request.form.get("question_type")
    if text:
        question.text = text
    if question_type in ("scale", "open"):
        question.question_type = question_type
    db.session.commit()
    return redirect(url_for("admin.admin_questions"))


@admin_bp.route("/questions/delete/<int:id>", methods=["POST"])
@admin_required
def admin_delete_question(id):
    question = Question.query.get_or_404(id)
    if question.responses:
        flash(
            "Deze vraag kan niet verwijderd worden omdat er al antwoorden op gegeven zijn.",
            "error",
        )
    else:
        db.session.delete(question)
        db.session.commit()
    return redirect(url_for("admin.admin_questions"))


@admin_bp.route("/tokens")
@admin_required
def admin_tokens():
    tokens = Token.query.order_by(Token.created_at.desc()).all()
    return render_template("admin_tokens.html", tokens=tokens)


@admin_bp.route("/tokens/generate", methods=["POST"])
@admin_required
def admin_generate_token():
    tokens = []
    token = Token(token=Token.generate_token())
    db.session.add(token)
    tokens.append(token)
    db.session.commit()
    return redirect(url_for("admin.admin_tokens"))


@admin_bp.route("/tokens/delete/<int:id>", methods=["POST"])
@admin_required
def admin_delete_token(id):
    token = Token.query.get_or_404(id)

    if token.used and token.responses:
        flash(
            "Deze uitnodiging kan niet verwijderd worden omdat de vragenlijst al is ingevuld.",
            "Verwijder eerst de bijbehorende antwoorden als u deze uitnodiging wilt verwijderen.",
            "error",
        )
    else:
        db.session.delete(token)
        db.session.commit()
    return redirect(url_for("admin.admin_tokens"))


@admin_bp.route("/responses")
@admin_required
def admin_responses():
    return redirect(url_for("admin.admin_dashboard"))


@admin_bp.route("/export")
@admin_required
def admin_export_responses():
    wb = Workbook()

    # Stijlen
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Tab 1: Gemiddelde
    ws_avg = wb.active
    ws_avg.title = "Gemiddelde"

    questions = Question.query.order_by(Question.order).all()

    ws_avg["A1"] = "Nr."
    ws_avg["B1"] = "Vraag"
    ws_avg["C1"] = "Gemiddelde Score"
    ws_avg["D1"] = "Aantal Antwoorden"

    for col in ["A", "B", "C", "D"]:
        cell = ws_avg[f"{col}1"]
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment

    row = 2
    for idx, question in enumerate(questions, 1):
        ws_avg[f"A{row}"] = idx
        ws_avg[f"B{row}"] = question.text

        if question.question_type == "scale":
            scores = [
                r.score
                for r in Response.query.filter_by(question_id=question.id).all()
                if r.score is not None
            ]
            ws_avg[f"C{row}"] = round(sum(scores) / len(scores), 1) if scores else 0
            ws_avg[f"D{row}"] = len(scores)
        else:
            text_answers = [
                r.text_answer
                for r in Response.query.filter_by(question_id=question.id).all()
                if r.text_answer
            ]
            ws_avg[f"C{row}"] = "Open vraag"
            ws_avg[f"D{row}"] = len(text_answers)

        for col in ["A", "B", "C", "D"]:
            ws_avg[f"{col}{row}"].border = border

        ws_avg[f"A{row}"].alignment = center_alignment
        ws_avg[f"C{row}"].alignment = center_alignment
        ws_avg[f"D{row}"].alignment = center_alignment

        row += 1

    ws_avg.column_dimensions["A"].width = 6
    ws_avg.column_dimensions["B"].width = 60
    ws_avg.column_dimensions["C"].width = 18
    ws_avg.column_dimensions["D"].width = 20

    # Tab 2: Individueel
    ws_ind = wb.create_sheet(title="Individueel")

    responses = (
        db.session.query(
            Response.token_id,
            Response.submitted_at,
            Question.order,
            Question.text,
            Question.question_type,
            Response.score,
            Response.text_answer,
        )
        .join(Question)
        .order_by(Response.token_id, Question.order)
        .all()
    )

    grouped_responses = {}
    for response in responses:
        token_id = response.token_id
        if token_id not in grouped_responses:
            grouped_responses[token_id] = {
                "submitted_at": response.submitted_at,
                "answers": {},
            }
        if response.question_type == "open":
            grouped_responses[token_id]["answers"][response.order] = response.text_answer or ""
        else:
            grouped_responses[token_id]["answers"][response.order] = response.score

    ws_ind["A1"] = "Werknemer ID"
    ws_ind["B1"] = "Ingediend op"

    col = 3
    for idx, question in enumerate(questions, 1):
        col_letter = get_column_letter(col)
        ws_ind[f"{col_letter}1"] = f"V{idx}"
        ws_ind[f"{col_letter}1"].fill = header_fill
        ws_ind[f"{col_letter}1"].font = header_font
        ws_ind[f"{col_letter}1"].border = border
        ws_ind[f"{col_letter}1"].alignment = center_alignment
        col += 1

    for col in ["A", "B"]:
        cell = ws_ind[f"{col}1"]
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment

    row = 2
    for token_id, data in sorted(grouped_responses.items()):
        ws_ind[f"A{row}"] = token_id
        ws_ind[f"B{row}"] = data["submitted_at"].strftime("%d-%m-%Y %H:%M")

        col = 3
        for question in questions:
            col_letter = get_column_letter(col)
            score = data["answers"].get(question.order, "")
            ws_ind[f"{col_letter}{row}"] = score
            ws_ind[f"{col_letter}{row}"].alignment = center_alignment
            ws_ind[f"{col_letter}{row}"].border = border
            col += 1

        ws_ind[f"A{row}"].border = border
        ws_ind[f"B{row}"].border = border
        ws_ind[f"A{row}"].alignment = center_alignment

        row += 1

    ws_ind.column_dimensions["A"].width = 15
    ws_ind.column_dimensions["B"].width = 18
    for col_idx, question in enumerate(questions):
        col_num = 3 + col_idx
        if question.question_type == "open":
            ws_ind.column_dimensions[get_column_letter(col_num)].width = 40
        else:
            ws_ind.column_dimensions[get_column_letter(col_num)].width = 8

    # Vraag referentie
    if questions:
        note_row = row + 2
        ws_ind[f"A{note_row}"] = "Vraag Referentie:"
        ws_ind[f"A{note_row}"].font = Font(bold=True)
        note_row += 1
        for idx, question in enumerate(questions, 1):
            ws_ind[f"A{note_row}"] = f"V{idx}:"
            ws_ind[f"B{note_row}"] = question.text
            ws_ind.merge_cells(f"B{note_row}:Z{note_row}")
            note_row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = (
        f'attachment; filename=antwoorden_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

    return response


@admin_bp.route("/reset", methods=["GET", "POST"])
@admin_required
def admin_reset():
    if request.method == "POST":
        try:
            Response.query.delete()
            Token.query.delete()

            db.session.commit()
            return redirect(url_for("admin.admin_dashboard"))
        except Exception as e:
            db.session.rollback()
            flash(f"Er is een fout opgetreden: {str(e)}", "error")
            return redirect(url_for("admin.admin_reset"))

    has_responses = Response.query.count() > 0
    return render_template("admin_reset.html", has_responses=has_responses)
